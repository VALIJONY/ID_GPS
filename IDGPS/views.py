from django.urls import reverse_lazy
import jwt
from datetime import datetime, timedelta
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate
from django.contrib.auth.views import LoginView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import login,logout
from .models import Bugalteriya, CustomUser, Note, Rasxod,Sklad, Sotish, DasturiyTaminot, MashinaMalumoti
from django.views import View
from .forms import SkladForm,SotishForm, HodimForm
from django.views.generic import UpdateView,DeleteView,TemplateView,ListView,CreateView
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponseBadRequest, JsonResponse
from django.db.models import Q
from itertools import zip_longest
from django.db import models
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse

class Loginview(LoginView):
    template_name = 'registration/login.html'  

    def get_success_url(self):
        return reverse_lazy('home')

    def form_valid(self, form):
        user = form.get_user()
        login(self.request, user)

        payload = {
            'id': user.id,
            'username': user.username,
            'exp': datetime.utcnow() + timedelta(hours=1),  
            'iat': datetime.utcnow(), 
        }
        access_token = jwt.encode(payload, settings.SECRET_KEY, algorithm='HS256')

        
        response = redirect(self.get_success_url()) 
        response.set_cookie('access_token', access_token)
        response.set_cookie('message', 'Success')  
        return response

    def form_invalid(self, form):
        return self.render_to_response(
            self.get_context_data(
                form=form,
                message="Invalid username or password"
            )
        )

class LogoutView(LoginRequiredMixin, View):
    def get(self, request):
        logout(request)
        response = redirect('login') 
        response.delete_cookie('access_token') 
        return response


class Home(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'base.html', {'user': request.user})
    
    
class SkladView(LoginRequiredMixin, View):
    def get(self,request):
        sklad_list=Sklad.objects.all().order_by('-olingan_sana')
        return render(request,'sklad.html',{'skladlist':sklad_list})
    
    
class SkladAddView(LoginRequiredMixin, View):
    template_name = 'sklad.html'
    def get(self, request):
        form = SkladForm()
        return render(request, self.template_name, {'form': form})
    
    def post(self, request):
        form = SkladForm(request.POST)
        if form.is_valid():
            # Asosiy ma'lumotlarni olish
            olingan_odam = form.cleaned_data['olingan_odam']
            tel_raqam = form.cleaned_data['tel_raqam']
            summa_prixod = form.cleaned_data['summa_prixod']
            olingan_sana = form.cleaned_data['olingan_sana']
            
            # Barcha GPS ID larni POST dan olish
            gps_ids = []
            i = 1
            while True:
                gps_id = request.POST.get(f'gps_id_{i}')
                if not gps_id and i == 1:
                    # Birinchi GPS ID formadan olish
                    gps_id = form.cleaned_data.get('gps_id')
                if not gps_id:
                    break
                gps_ids.append(gps_id)
                i += 1
            
            # Har bir GPS ID uchun alohida yozuv yaratish
            for gps_id in gps_ids:
                Sklad.objects.create(
                    gps_id=gps_id,
                    olingan_odam=olingan_odam,
                    tel_raqam=tel_raqam,
                    summa_prixod=summa_prixod,
                    olingan_sana=olingan_sana,
                    sotildi_sotilmadi=False
                )
            
            messages.success(request, "GPS ma'lumotlari muvaffaqiyatli saqlandi!")
            return redirect('sklad-list')
            
        return render(request, self.template_name, {'form': form})
    

class SkladUpdateView(LoginRequiredMixin, View):
    def get(self, request, pk):
        sklad = Sklad.objects.get(pk=pk)
        form = SkladForm(instance=sklad)
        return render(request, 'sklad_update.html', {'form': form})

    def post(self, request, pk):
        sklad = Sklad.objects.get(pk=pk)
        form = SkladForm(request.POST, instance=sklad)
        if form.is_valid():
            form.save()
            return redirect('sklad-list')
        return render(request, 'sklad_update.html', {'form': form})

class SkladDeleteView(LoginRequiredMixin, View):
    def get(self, request, pk):
        sklad = Sklad.objects.get(pk=pk)
        sklad.delete()
        return redirect('sklad-list')


def sklad_list(request):
    status_filter = request.GET.get('status', '')
    skladlist = Sklad.objects.all()

    if status_filter == 'sold':
        skladlist = skladlist.filter(sotildi_sotilmadi=True)
    elif status_filter == 'unsold':
        skladlist = skladlist.filter(sotildi_sotilmadi=False)
        
    skladlist = skladlist.order_by('-olingan_sana')
    return render(request, 'sklad.html', {'skladlist': skladlist})


class SotishListView(LoginRequiredMixin, View):
    def get(self, request):
        sotish_items = Sotish.objects.all().order_by('-sana')  # Sana bo'yicha tartiblangan ro'yxat
        return render(request, 'sotish.html', {'sotish_items': sotish_items})

class SotishAddView(View):
    template_name = 'sotish.html'

    def get(self, request):
        form = SotishForm()
        gps_list = Sklad.objects.filter(sotildi_sotilmadi=False)  # Sotilmagan GPS larni olish
        return render(request, self.template_name, {
            'form': form,
            'gps_list': gps_list
        })

    def post(self, request):
        form = SotishForm(request.POST)
        
        if form.is_valid():
            try:
                # Raqamlarni formatlash va o'zgartirish
                def format_number(value):
                    if not value:
                        return 0
                    # Vergul va probellarni olib tashlash
                    value = value.replace(',', '').replace(' ', '')
                    try:
                        return int(value)
                    except ValueError:
                        return 0

                summasi = format_number(request.POST.get('summasi', '0'))
                naqd = format_number(request.POST.get('naqd', '0'))
                karta = format_number(request.POST.get('karta', '0'))
                bank_schot = format_number(request.POST.get('bank_schot', '0'))
                master_summasi = format_number(request.POST.get('master_summasi', '0'))

                # Umumiy summa va qarz hisoblash
                umumiy_summasi = summasi
                qarz = umumiy_summasi - (naqd + bank_schot)
                
                if qarz < 0:
                    messages.error(request, "To'lov summasi umumiy summadan oshib ketdi!")
                    return render(request, self.template_name, {
                        'form': form,
                        'gps_list': Sklad.objects.filter(sotildi_sotilmadi=False)
                    })

                # Sotish obyektini yaratish
                sotish = form.save(commit=False)
                sotish.summasi = summasi
                sotish.naqd = naqd
                sotish.karta = karta
                sotish.bank_schot = bank_schot
                sotish.master_summasi = master_summasi
                sotish.save()

                # GPS va mashina ma'lumotlarini olish
                gps_ids = form.cleaned_data['gps_id']
                sim_kartalar = request.POST.getlist('sim_karta')
                mashina_turlari = request.POST.getlist('mashina_turi')
                davlat_raqamlari = request.POST.getlist('davlat_raqami')

                print("Kelgan ma'lumotlar:")
                print("GPS IDs:", gps_ids)
                print("SIM kartalar:", sim_kartalar)
                print("Mashina turlari:", mashina_turlari)
                print("Davlat raqamlari:", davlat_raqamlari)

                # SIM kartalarni vergul bilan ajratib saqlash
                sotish.sim_karta = ', '.join(filter(None, sim_kartalar))
                sotish.save()

                # GPS va mashina ma'lumotlarini saqlash
                for i, gps in enumerate(gps_ids):
                    gps.sotildi_sotilmadi = True
                    gps.save()
                    sotish.gps_id.add(gps)
                    
                    # Mashina ma'lumotlarini saqlash
                    if i < len(mashina_turlari) and i < len(davlat_raqamlari):
                        MashinaMalumoti.objects.create(
                            sotish=sotish,
                            gps=gps,
                            mashina_turi=mashina_turlari[i],
                            davlat_raqami=davlat_raqamlari[i]
                        )
                        print(f"Mashina ma'lumotlari saqlandi: {mashina_turlari[i]} - {davlat_raqamlari[i]}")

                messages.success(request, "Sotish muvaffaqiyatli saqlandi!")
                return redirect('sotish_list')
            
            except Exception as e:
                print("Asosiy xatolik:", str(e))
                messages.error(request, f"Xatolik yuz berdi: {str(e)}")
        else:
            print("Forma xatoliklari:", form.errors)
            messages.error(request, "Forma to'g'ri to'ldirilmagan!")
        
        return render(request, self.template_name, {
            'form': form,
            'gps_list': Sklad.objects.filter(sotildi_sotilmadi=False)
        })

class SotishUpdateView(LoginRequiredMixin, View):
    def get(self, request, pk):
        sotish = get_object_or_404(Sotish, id=pk)
        form = SotishForm(instance=sotish)
        
        # GPS, SIM karta va mashina ma'lumotlarini olish
        sim_kartalar = sotish.sim_karta.split(', ') if sotish.sim_karta else []
        mashina_malumotlari = sotish.mashina_malumotlari.all()
        gps_data = []
        
        # GPS va unga tegishli ma'lumotlarni yig'ish
        for i, gps in enumerate(sotish.gps_id.all()):
            sim_karta = sim_kartalar[i] if i < len(sim_kartalar) else ""
            mashina = mashina_malumotlari.filter(gps=gps).first()
            
            gps_data.append({
                'gps': gps,
                'sim': sim_karta,
                'mashina_turi': mashina.mashina_turi if mashina else "",
                'davlat_raqami': mashina.davlat_raqami if mashina else ""
            })
        
        context = {
            'form': form,
            'sotish': sotish,
            'gps_data': gps_data,
            'mavjud_gpslar': Sklad.objects.filter(Q(sotildi_sotilmadi=False) | Q(id__in=sotish.gps_id.all()))
        }
        return render(request, 'sotish_update.html', context)

    def post(self, request, pk):
        sotish = get_object_or_404(Sotish, id=pk)
        form = SotishForm(request.POST, instance=sotish)
        
        try:
            # GPS va mashina ma'lumotlarini olish
            gps_ids = request.POST.getlist('gps_id')
            sim_kartalar = request.POST.getlist('sim_karta')
            mashina_turlari = request.POST.getlist('mashina_turi')
            davlat_raqamlari = request.POST.getlist('davlat_raqami')

            print("Kelgan ma'lumotlar:")
            print("GPS IDs:", gps_ids)
            print("SIM kartalar:", sim_kartalar)
            print("Mashina turlari:", mashina_turlari)
            print("Davlat raqamlari:", davlat_raqamlari)

            if form.is_valid():
                # Eski GPS va mashina ma'lumotlarini o'chirish
                old_gps_list = list(sotish.gps_id.all())
                print("Eski GPS lar:", [g.id for g in old_gps_list])
                
                for old_gps in old_gps_list:
                    old_gps.sotildi_sotilmadi = False
                    old_gps.save()
                
                sotish.gps_id.clear()
                sotish.mashina_malumotlari.all().delete()

                # Sotish obyektini yangilash
                sotish = form.save(commit=False)

                def format_number(value):
                    if not value:
                        return 0
                    # Vergul va probellarni olib tashlash
                    value = value.replace(',', '').replace(' ', '')
                    try:
                        return int(value)
                    except ValueError:
                        return 0

                sotish.summasi = format_number(request.POST.get('summasi', '0'))
                sotish.naqd = format_number(request.POST.get('naqd', '0'))
                sotish.karta = format_number(request.POST.get('karta', '0'))
                sotish.bank_schot = format_number(request.POST.get('bank_schot', '0'))
                sotish.master_summasi = format_number(request.POST.get('master_summasi', '0'))

                # SIM kartalarni vergul bilan ajratib saqlash
                sotish.sim_karta = ', '.join(filter(None, sim_kartalar))
                sotish.save()

                # Yangi GPS va mashina ma'lumotlarini saqlash
                for i in range(len(gps_ids)):
                    try:
                        gps_id = gps_ids[i]
                        if gps_id:  # Agar GPS ID bo'sh bo'lmasa
                            print(f"GPS {i} - ID: {gps_id}")
                            gps = Sklad.objects.get(id=gps_id)
                            gps.sotildi_sotilmadi = True
                            gps.save()
                            sotish.gps_id.add(gps)
                            
                            # Mashina ma'lumotlarini saqlash
                            if i < len(mashina_turlari) and i < len(davlat_raqamlari):
                                MashinaMalumoti.objects.create(
                                    sotish=sotish,
                                    gps=gps,
                                    mashina_turi=mashina_turlari[i],
                                    davlat_raqami=davlat_raqamlari[i]
                                )
                                print(f"Mashina ma'lumotlari saqlandi: {mashina_turlari[i]} - {davlat_raqamlari[i]}")
                    except Exception as e:
                        print(f"GPS {i} uchun xatolik: {str(e)}")

                messages.success(request, "Sotish muvaffaqiyatli yangilandi!")
                return redirect('sotish_list')
            else:
                print("Forma xatoliklari:", form.errors)
                messages.error(request, "Forma to'g'ri to'ldirilmagan!")
        
        except Exception as e:
            print("Asosiy xatolik:", str(e))
            messages.error(request, f"Xatolik yuz berdi: {str(e)}")
        
        return render(request, 'sotish_update.html', {
            'form': form,
            'sotish': sotish,
            'mavjud_gpslar': Sklad.objects.filter(Q(sotildi_sotilmadi=False) | Q(id__in=sotish.gps_id.all()))
        })

class SotishDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            sotish = get_object_or_404(Sotish, pk=pk)
            
            # Sklad bilan bog'liq GPSlarni qayta "sotilmadi" holatiga o'zgartirish
            gpslar = sotish.gps_id.all()
            for gps in gpslar:
                gps.sotildi_sotilmadi = False
                gps.save()
            
            # Mashina ma'lumotlarini o'chirish
            sotish.mashina_malumotlari.all().delete()
            
            # Sotishni o'chirish
            sotish.delete()
            messages.success(request, "Sotish muvaffaqiyatli o'chirildi!")
        except Exception as e:
            print("Sotishni o'chirishda xatolik:", str(e))
            messages.error(request, f"Sotishni o'chirishda xatolik yuz berdi: {str(e)}")
        
        return redirect('sotish_list')



class RasxodListView(LoginRequiredMixin, View):
    def get(self, request):
        rasxodlar = Rasxod.objects.all()
        return render(request, 'rasxod.html', {'rasxod_list': rasxodlar})
      
class RasxodAddView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'rasxod.html')

    def post(self, request):
        rasxod_nomi = request.POST.get('rasxod_nomi')
        sana = request.POST.get('sana')
        summa = request.POST.get('summasi')

        # Rasxodni yaratish
        Rasxod.objects.create(
            rasxod_nomi=rasxod_nomi,
            sana=sana,
            summa=summa
        )

        return redirect('rasxod_list')
    
class RasxodUpdateView(LoginRequiredMixin, UpdateView):
    model = Rasxod
    fields = ['rasxod_nomi', 'sana', 'summa']
    template_name = 'rasxod_update.html'  
    success_url = reverse_lazy('rasxod_list')  
    
class RasxodDeleteView(LoginRequiredMixin, View):
    def get(self, request, pk):
        rasxod = Rasxod.objects.get(pk=pk)
        rasxod.delete()
        return redirect('rasxod_list')
    
class MijozlarView(LoginRequiredMixin, TemplateView):
    template_name = 'mijozlar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mijozlar = Sotish.objects.all().order_by('-sana')  # Sanaga ko'ra tartiblash

        mijozlar_data = []
        for mijoz in mijozlar:
            gpslar = mijoz.gps_id.all()
            sim_kartalar = mijoz.sim_karta.split(',') if mijoz.sim_karta else []
            
            # GPS va SIM kartalar uzunligini tenglashtirish
            while len(sim_kartalar) < len(gpslar):
                sim_kartalar.append('')
            
            # Har bir GPS va SIM karta uchun qator yaratish
            for i, (gps, sim) in enumerate(zip(gpslar, sim_kartalar)):
                # GPS ga tegishli mashina ma'lumotlarini olish
                mashina = mijoz.mashina_malumotlari.filter(gps=gps).first()
                jami_summa=mijoz.summasi+mijoz.master_summasi+mijoz.abonent_tulov*mijoz.gps_id.count()
                mijozlar_data.append({
                    'id': mijoz.id,  # Mijoz ID si
                    'mijoz': mijoz.mijoz if i == 0 else '',
                    'tel_raqam': mijoz.mijoz_tel_raqam if i == 0 else '',
                    'username': mijoz.username if i == 0 else '',
                    'password': mijoz.password if i == 0 else '',
                    'dasturiy_taminot_nomi': str(mijoz.dasturiy_taminot) if mijoz.dasturiy_taminot and i == 0 else '',
                    'abonent_tulov': mijoz.abonent_tulov if i == 0 else '',
                    'sana': mijoz.sana.strftime('%Y-%m-%d') if mijoz.sana and i == 0 else '',
                    'summasi': mijoz.summasi if i == 0 else '',
                    'naqd': mijoz.naqd if i == 0 else '',
                    'bank_schot': mijoz.bank_schot if i == 0 else '',
                    'karta': mijoz.karta if i == 0 else '',
                    'master': mijoz.master if i == 0 else '',
                    'master_summasi': mijoz.master_summasi if i == 0 else '',
                    'gps': gps,
                    'sim_karta': sim.strip() if sim else '',  # Probellarni olib tashlash
                    'mashina_turi': mashina.mashina_turi if mashina else '',  # Mashina turi
                    'davlat_raqami': mashina.davlat_raqami if mashina else '',  # Davlat raqami
                    'rowspan': len(gpslar) if i == 0 else 0,
                    'first_row': i == 0,  # Birinchi qator uchun belgi
                    'jami_summa': jami_summa
                })

        context['mijozlar_data'] = mijozlar_data
        return context

class StatistikaView(LoginRequiredMixin, TemplateView):
    template_name = 'statistika.html'
    oylar = ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
             "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]

    def get_stats(self, month, year):
        """Berilgan oy va yil uchun statistikani hisoblash"""
        # O'tgan oy va yilni hisoblash
        prev_month = 12 if month == 1 else month - 1
        prev_year = year - 1 if month == 1 else year

        # Abonentlar statistikasi
        abonent_stats = {
            'oldingidagi_aktiv': Sotish.objects.filter(
                sana__year=prev_year,
                sana__month=prev_month
            ).aggregate(gps_count=models.Count('gps_id'))['gps_count'],
            'oylik_qoshilgan': Sotish.objects.filter(
                sana__year=year,
                sana__month=month
            ).aggregate(gps_count=models.Count('gps_id'))['gps_count'],
            'jami_aktiv': Sotish.objects.filter(
                models.Q(sana__year__lt=year) |  # o'tgan yillarning hammasi
                models.Q(sana__year=year, sana__month__lte=month)  # joriy yilning belgilangan oyigacha
            ).aggregate(gps_count=models.Count('gps_id'))['gps_count']
        }

        # To'lovlar statistikasi
        tolov_stats = {
            'oylik_tolamaganlar': Bugalteriya.objects.filter(
                yil=year,
                oy=self.oylar[month-1],
                abonent_tolov=False
            ).aggregate(gps_count=models.Count('gps_id'))['gps_count'],
            'oylik_tolaganlar': Bugalteriya.objects.filter(
                yil=year,
                oy=self.oylar[month-1],
                abonent_tolov=True
            ).aggregate(gps_count=models.Count('gps_id'))['gps_count'],
        }
        print(tolov_stats)
        # GPS qurilmalari statistikasi
        gps_stats = {
            'otgan_oy_sklad': Sklad.objects.filter(
                olingan_sana__year=prev_year,
                olingan_sana__month=prev_month,
                sotildi_sotilmadi=False
            ).count(),
            'hozir_skladda_bor': Sklad.objects.filter(
                sotildi_sotilmadi=False
            ).count(),
            'oylik_sotilgan': Sklad.objects.filter(
                sotildi_sotilmadi=True,
                sotish__sana__year=year,
                sotish__sana__month=month
            ).count(),
            'jami_sotilgan': Sklad.objects.filter(
                sotildi_sotilmadi=True,
                sotish__sana__year__lte=year,
                sotish__sana__month__lte=month if year == datetime.now().year else 12
            ).count()
        }

        # SIM karta statistikasi
        sim_stats = {
            'oylik': Sotish.objects.filter(
                sana__year=year,
                sana__month=month
            ).annotate(
                sim_count=models.Func(
                    models.F('sim_karta'),
                    function='LENGTH',
                    template="(LENGTH(%(expressions)s) - LENGTH(REPLACE(%(expressions)s, ',', '')) + 1)"
                )
            ).aggregate(total_sim=models.Sum('sim_count'))['total_sim'] or 0
        }

        # Summalar statistikasi    
        summa_stats = {
            'oylik_umumiy_summa': Sotish.objects.filter(
                sana__year=year,
                sana__month=month
            ).aggregate(total=models.Sum('summasi'))['total'] or 0,
            'oylik_abonent': Bugalteriya.objects.filter(
                yil=year,
                oy=self.oylar[month-1],
                abonent_tolov=True
            ).aggregate(total=models.Sum('sotish__abonent_tulov'))['total'] or 0
        }

        return {
            'abonent': abonent_stats,
            'tolov': tolov_stats,
            'gps': gps_stats,
            'sim': sim_stats,
            'summa': summa_stats,
            'oy': self.oylar[month-1]
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        now = datetime.now()
        selected_year = int(self.request.GET.get('year', now.year))
        
        # Har bir oy uchun statistika olish
        monthly_stats = {}
        current_month = now.month if selected_year == now.year else 12
        
        for month in range(1, current_month + 1):
            monthly_stats[month] = self.get_stats(month, selected_year)
        
        context.update({
            'monthly_stats': monthly_stats,
            'current_year': selected_year,
            'current_month': now.month,
            'yillar': range(2020, now.year + 1)
        })
        
        return context

class BugalteriyaView(LoginRequiredMixin, TemplateView):
    template_name = 'bugalteriya.html'
    oylar = ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
             "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        now = datetime.now()
        selected_year = int(self.request.GET.get('yil', now.year))
        
        sotishlar = (Sotish.objects
               .filter(sana__year__lte=selected_year)
               .prefetch_related('gps_id')
               .distinct()
               .order_by('-sana'))

        tolovlar = (Bugalteriya.objects
                 .filter(sotish__in=sotishlar)
                 .filter(yil=selected_year)
                 .select_related('sotish', 'gps'))
        
        tolovlar_dict = {}
        for tolov in tolovlar:
            key = (tolov.sotish_id, tolov.gps_id, tolov.oy)
            tolovlar_dict[key] = tolov
        
        bugalteriya_data = []
        for sotish in sotishlar:
            gps_data = []
            for gps in sotish.gps_id.all():
                oylik_tolovlar = {}
                for oy in self.oylar:
                    tolov = tolovlar_dict.get((sotish.id, gps.id, oy))
                    oylik_tolovlar[oy] = {
                        'abonent': {
                            'status': tolov.abonent_tolov if tolov else None,  # None bo'lsa "To'lovni tasdiqlash" ko'rinadi
                            'id': tolov.id if tolov else '',
                        },
                        'sim': {
                            'status': tolov.sim_karta_tolov if tolov else None,  # None bo'lsa "To'lovni tasdiqlash" ko'rinadi
                            'id': tolov.id if tolov else '',
                        },
                        'tolov': tolov  # Qo'shimcha maydon - tolov obyekti mavjudligini tekshirish uchun
                    }
                
                gps_data.append({
                    'gps': gps,
                    'tolovlar': oylik_tolovlar
                })
            
            bugalteriya_data.append({
                'sotish': sotish,
                'gps_data': gps_data
            })
        
        context.update({
            'bugalteriya_data': bugalteriya_data,
            'oylar': self.oylar,
            'current_year': selected_year,
            'years': range(2020, now.year + 1),
            'is_superuser': self.request.user.is_superuser
        })
        
        return context

class UpdateBugalteriyaView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            tolov_id = data.get('tolov_id')
            sotish_id = data.get('sotish_id')
            gps_id = data.get('gps_id')
            oy = data.get('oy')
            yil = data.get('yil')
            tolov_type = data.get('type')

            # Agar to'lov mavjud bo'lsa
            if tolov_id and tolov_id != 'null' and tolov_id != 'undefined':
                tolov = get_object_or_404(Bugalteriya, id=tolov_id)
                
                if tolov_type == 'abonent':
                    current_status = tolov.abonent_tolov
                    # Oddiy foydalanuvchi uchun: False -> True -> None
                    # Superuser uchun: False -> True -> None -> False
                    if current_status is False:
                        tolov.abonent_tolov = True
                    elif current_status is True:
                        tolov.abonent_tolov = None
                    elif current_status is None and request.user.is_superuser:
                        tolov.abonent_tolov = False
                    else:
                        return JsonResponse({
                            'status': False,
                            'message': "Null holatdagi to'lovni faqat super admin o'zgartira oladi"
                        })
                else:  # sim karta to'lovi
                    current_status = tolov.sim_karta_tolov
                    if current_status is False:
                        tolov.sim_karta_tolov = True
                    elif current_status is True:
                        tolov.sim_karta_tolov = None
                    elif current_status is None and request.user.is_superuser:
                        tolov.sim_karta_tolov = False
                    else:
                        return JsonResponse({
                            'status': False,
                            'message': "Null holatdagi to'lovni faqat super admin o'zgartira oladi"
                        })
            
            # Yangi to'lov yaratish
            else:
                if not all([sotish_id, gps_id, oy, yil]):
                    return JsonResponse({
                        'status': False,
                        'message': "To'lov uchun barcha ma'lumotlar to'liq emas"
                    })

                tolov = Bugalteriya.objects.filter(
                    sotish_id=sotish_id,
                    gps_id=gps_id,
                    oy=oy,
                    yil=yil
                ).first()

                if not tolov:
                    sotish = get_object_or_404(Sotish, id=sotish_id)
                    gps = get_object_or_404(Sklad, id=gps_id)
                    tolov = Bugalteriya.objects.create(
                        sotish=sotish,
                        gps=gps,
                        oy=oy,
                        yil=yil,
                        abonent_tolov=False,  # Default False
                        sim_karta_tolov=False
                    )

            # To'lovni saqlash
            tolov.save()

            return JsonResponse({
                'status': True,
                'tolov_id': tolov.id,
                'abonent_status': tolov.abonent_tolov,
                'sim_status': tolov.sim_karta_tolov,
            })

        except Exception as e:
            return JsonResponse({
                'status': False,
                'message': str(e)
            })
            
class HodimListView(LoginRequiredMixin, ListView):
    model = CustomUser
    template_name = 'hodim.html'
    context_object_name = 'hodimlar'

    def get_queryset(self):
        queryset = super().get_queryset()
        for user in queryset:
            # raw_password ni user.password ga o'rnatamiz
            user.raw_password = user._password
        return queryset
    
class HodimCreateView(LoginRequiredMixin, CreateView):
    model = CustomUser
    form_class = HodimForm
    template_name = 'hodim.html'
    success_url = reverse_lazy('hodim-list')

class HodimUpdateView(LoginRequiredMixin, UpdateView):
    model = CustomUser
    form_class = HodimForm
    template_name = 'hodim_update.html'
    success_url = reverse_lazy('hodim-list')

class HodimDeleteView(LoginRequiredMixin, View):
    def get(self, request, pk):
        hodim = get_object_or_404(CustomUser, pk=pk)
        hodim.delete()
        return redirect('hodim-list')
    
class NoteView(LoginRequiredMixin, View):
    def get(self, request):
        notes = Note.objects.all().order_by('-sana')
        return render(request, 'note.html', {'notes': notes})

class NoteAddView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'note.html') 
    
    def post(self, request):
        user = request.user
        note_content = request.POST.get('note')  
        if note_content:  
            note = Note(user=user, izoh=note_content)
            note.save() 
            return redirect('note-list') 
        return render(request, 'note.html', {'error': 'Please enter a note!'}) 

class NoteEditView(LoginRequiredMixin, View):
    def get(self, request, pk):
        note = get_object_or_404(Note, pk=pk)
        if note.user != request.user:
            return redirect('note-list')
        return render(request, 'note.html', {'note': note, 'is_edit': True})

    def post(self, request, pk):
        note = get_object_or_404(Note, pk=pk)
        if note.user != request.user:
            return redirect('note-list')
        note_content = request.POST.get('note')
        if note_content:
            note.izoh = note_content
            note.save()
            return redirect('note-list')
        return render(request, 'note.html', {'error': 'Please enter a note!', 'note': note, 'is_edit': True})

class NoteDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        note = get_object_or_404(Note, pk=pk)
        if note.user != request.user:
            return redirect('note-list')
        note.delete()
        return redirect('note-list')


class GPSAddExcelView(LoginRequiredMixin, View):
    template_name = 'sklad.html'

    def get(self, request):
        # Excel shablonini yaratish
        if request.GET.get('download_template'):
            # Excel fayl yaratish
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "GPS Ma'lumotlari"

            # Ustun nomlari
            headers = ['gps_id', 'olingan_odam', 'tel_raqam', 'summa_prixod', 'olingan_sana']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Namuna ma'lumotlar
            sample_data = [
                ['GPS001', 'Ali Valiyev', '+998901234567', 100000, '2025-01-27'],
                ['GPS002', 'Vali Aliyev', '+998907654321', 150000, '2025-01-27']
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value

            # Ustun kengliklari
            ws.column_dimensions['A'].width = 15  # gps_id
            ws.column_dimensions['B'].width = 25  # olingan_odam
            ws.column_dimensions['C'].width = 15  # tel_raqam
            ws.column_dimensions['D'].width = 15  # summa_prixod
            ws.column_dimensions['E'].width = 15  # olingan_sana

            # Excel faylni saqlash va yuborish
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=gps_template.xlsx'
            wb.save(response)
            return response

        return render(request, self.template_name, {'skladlist': Sklad.objects.all()})

    def post(self, request):
        try:
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "Excel fayl tanlanmagan!")
                return redirect('sklad-list')

            # Excel faylni tekshirish
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "Noto'g'ri fayl formati. Faqat .xlsx yoki .xls fayllar qabul qilinadi!")
                return redirect('sklad-list')

            # Excel faylni o'qish
            df = pd.read_excel(excel_file)
            required_columns = ['gps_id', 'olingan_odam', 'tel_raqam', 'summa_prixod', 'olingan_sana']
            
            # Ustunlar mavjudligini tekshirish
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messages.error(request, f"Excel faylda quyidagi ustunlar mavjud emas: {', '.join(missing_columns)}")
                return redirect('sklad-list')

            success_count = 0
            error_count = 0
            errors = []

            # Excel fayldagi ma'lumotlarni bazaga saqlash
            for index, row in df.iterrows():
                try:
                    # Bo'sh qatorlarni o'tkazib yuborish
                    if pd.isna(row['gps_id']) or str(row['gps_id']).strip() == '':
                        continue

                    # GPS ID ni tekshirish
                    gps_id = str(row['gps_id']).strip()
                    if Sklad.objects.filter(gps_id=gps_id).exists():
                        errors.append(f"GPS ID {gps_id} allaqachon mavjud!")
                        error_count += 1
                        continue

                    # Ma'lumotlarni tozalash
                    olingan_odam = str(row['olingan_odam']).strip() if not pd.isna(row['olingan_odam']) else ''
                    tel_raqam = str(row['tel_raqam']).strip() if not pd.isna(row['tel_raqam']) else ''
                    
                    try:
                        summa_prixod = float(row['summa_prixod'])
                    except (ValueError, TypeError):
                        summa_prixod = 0
                        
                    try:
                        olingan_sana = pd.to_datetime(row['olingan_sana']).date()
                    except:
                        olingan_sana = datetime.now().date()

                    # Yangi GPS qo'shish
                    Sklad.objects.create(
                        gps_id=gps_id,
                        olingan_odam=olingan_odam,
                        tel_raqam=tel_raqam,
                        summa_prixod=summa_prixod,
                        olingan_sana=olingan_sana,
                        sotildi_sotilmadi=False
                    )
                    success_count += 1

                except Exception as e:
                    errors.append(f"Qator {index + 2}: {str(e)}")
                    error_count += 1

            # Natijalarni xabar qilish
            if success_count > 0:
                messages.success(request, f"{success_count} ta GPS muvaffaqiyatli qo'shildi!")
            if error_count > 0:
                messages.error(request, f"{error_count} ta xatolik yuz berdi!")
                for error in errors:
                    messages.error(request, error)

        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {str(e)}")

        return redirect('sklad-list')